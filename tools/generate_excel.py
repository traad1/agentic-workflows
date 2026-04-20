#!/usr/bin/env python3
"""
Generate a 4-tab Excel ops report from .tmp/shipments.json and .tmp/changes.json.
Outputs .tmp/ops_report_YYYY-MM-DD.xlsx.

Usage: python3 tools/generate_excel.py
"""

import json
import sys
from datetime import datetime, date, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

SHIPMENTS_FILE = Path(__file__).parent.parent / ".tmp" / "shipments.json"
CHANGES_FILE = Path(__file__).parent.parent / ".tmp" / "changes.json"
OUTPUT_DIR = Path(__file__).parent.parent / ".tmp"

# ── Colours ────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_ROW_FILL = PatternFill("solid", fgColor="EEF2F7")
RED_FILL = PatternFill("solid", fgColor="FFCCCC")       # overdue / urgent
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")    # arriving soon / changed
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")     # new / resolved
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")    # removed / missing
NULL_FILL = PatternFill("solid", fgColor="FFFACD")      # missing fields

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header_style(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def data_style(cell, fill=None, bold=False):
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    cell.border = BORDER
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)


def auto_width(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def parse_date_loose(s):
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y", "%B %d %Y",
                "%b %d %Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except Exception:
            pass
    return None


def is_arriving_this_week(eta_str):
    d = parse_date_loose(eta_str)
    if not d:
        return False
    today = date.today()
    return today <= d <= today + timedelta(days=7)


def is_overdue_eta(eta_str):
    d = parse_date_loose(eta_str)
    if not d:
        return False
    return d < date.today()


# ── Tab 1: Active Shipments ─────────────────────────────────────────────────

def write_active_shipments(wb, shipments):
    ws = wb.create_sheet("Active Shipments")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = ["Vessel", "Origin", "Grade", "Qty MT", "Qty Bags",
               "BL Number", "ETD", "ETA US", "Dest Port",
               "Contract Ref", "Status", "Last Update", "Source"]
    fields = ["vessel", "origin", "grade", "quantity_mt", "quantity_bags",
              "bl_number", "etd", "eta_us", "dest_port",
              "contract_ref", "status", "last_update_date", "source_email"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)

    for row_idx, s in enumerate(shipments, 2):
        alt = ALT_ROW_FILL if row_idx % 2 == 0 else None
        eta = s.get("eta_us")
        row_fill = (RED_FILL if is_overdue_eta(eta)
                    else YELLOW_FILL if is_arriving_this_week(eta)
                    else alt)

        for col, field in enumerate(fields, 1):
            val = s.get(field)
            if field == "last_update_date" and val:
                val = val[:10]
            cell = ws.cell(row=row_idx, column=col, value=val)
            field_fill = NULL_FILL if val is None else row_fill
            data_style(cell, fill=field_fill)

    auto_width(ws)

    ws.cell(row=len(shipments) + 3, column=1,
            value="🔴 Red = ETA overdue   🟡 Yellow = arriving this week   🟡 Light yellow = missing field")


# ── Tab 2: Changes This Week ─────────────────────────────────────────────────

def write_changes(wb, changes, shipments_lookup):
    ws = wb.create_sheet("Changes This Week")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    if changes.get("baseline"):
        ws.cell(row=1, column=1, value="Baseline week — no previous data to compare against.")
        ws.cell(row=1, column=1).font = Font(italic=True, color="888888")
        return

    row = 1

    def section_header(label):
        nonlocal row
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, size=11, color="1F3864")
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

    def col_headers(*labels):
        nonlocal row
        for col, h in enumerate(labels, 1):
            cell = ws.cell(row=row, column=col, value=h)
            header_style(cell)
        row += 1

    # New shipments
    new_ships = changes.get("new_shipments", [])
    section_header(f"NEW SHIPMENTS ({len(new_ships)})")
    col_headers("Vessel", "Origin", "Grade", "Qty MT", "ETA US", "Status")
    for s in new_ships:
        for col, field in enumerate(["vessel", "origin", "grade", "quantity_mt", "eta_us", "status"], 1):
            cell = ws.cell(row=row, column=col, value=s.get(field))
            data_style(cell, fill=GREEN_FILL)
        row += 1
    if not new_ships:
        ws.cell(row=row, column=1, value="None").font = Font(italic=True, color="888888")
        row += 1
    row += 1

    # Updated shipments
    updated = changes.get("updated_shipments", [])
    section_header(f"UPDATED SHIPMENTS ({len(updated)})")
    col_headers("Vessel / BL", "Field", "Previous Value", "New Value", "", "")
    for u in updated:
        label = u.get("vessel") or u.get("bl_number") or "Unknown"
        for change in u.get("changes", []):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=change["field"])
            old_cell = ws.cell(row=row, column=3, value=str(change["old"]) if change["old"] is not None else "—")
            new_cell = ws.cell(row=row, column=4, value=str(change["new"]) if change["new"] is not None else "—")
            data_style(old_cell, fill=RED_FILL)
            data_style(new_cell, fill=GREEN_FILL)
            row += 1
    if not updated:
        ws.cell(row=row, column=1, value="No field changes").font = Font(italic=True, color="888888")
        row += 1
    row += 1

    # Removed shipments
    removed = changes.get("removed_shipments", [])
    section_header(f"REMOVED / NO LONGER REPORTED ({len(removed)})")
    col_headers("Vessel", "BL Number", "Origin", "Last ETA", "Last Status", "")
    for s in removed:
        for col, field in enumerate(["vessel", "bl_number", "origin", "eta_us", "status"], 1):
            cell = ws.cell(row=row, column=col, value=s.get(field))
            data_style(cell, fill=ORANGE_FILL)
        row += 1
    if not removed:
        ws.cell(row=row, column=1, value="None").font = Font(italic=True, color="888888")
        row += 1
    row += 1

    # Resolved subjects
    resolved = changes.get("resolved_subjects", [])
    section_header(f"RESOLVED PENDING SUBJECTS ({len(resolved)})")
    col_headers("Reference", "Subject Type", "Description", "Reported By", "", "")
    for s in resolved:
        for col, field in enumerate(["reference", "subject_type", "description", "reported_by"], 1):
            cell = ws.cell(row=row, column=col, value=s.get(field))
            data_style(cell, fill=GREEN_FILL)
        row += 1
    if not resolved:
        ws.cell(row=row, column=1, value="None").font = Font(italic=True, color="888888")
        row += 1

    auto_width(ws)


# ── Tab 2b: By Origin ────────────────────────────────────────────────────────

def write_by_origin(wb, shipments):
    ws = wb.create_sheet("By Origin")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # Group shipments by origin
    from collections import defaultdict
    by_origin = defaultdict(list)
    for s in shipments:
        origin = s.get("origin") or "Unknown"
        by_origin[origin].append(s)

    headers = ["Vessel", "Grade", "Qty MT", "Qty Bags", "BL Number", "ETD", "ETA US", "Dest Port", "Status"]
    fields = ["vessel", "grade", "quantity_mt", "quantity_bags", "bl_number", "etd", "eta_us", "dest_port", "status"]

    row = 1
    for origin in sorted(by_origin.keys()):
        group = by_origin[origin]

        # Origin section header
        origin_cell = ws.cell(row=row, column=1, value=f"  {origin.upper()}")
        origin_cell.font = Font(bold=True, color="FFFFFF", size=11)
        origin_cell.fill = PatternFill("solid", fgColor="2E4057")
        origin_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1

        # Column headers
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            header_style(cell)
        row += 1

        # Data rows
        total_mt = 0
        total_bags = 0
        for idx, s in enumerate(sorted(group, key=lambda x: x.get("eta_us") or "9999")):
            alt = ALT_ROW_FILL if idx % 2 == 0 else None
            eta = s.get("eta_us")
            row_fill = (RED_FILL if is_overdue_eta(eta)
                        else YELLOW_FILL if is_arriving_this_week(eta)
                        else alt)
            for col, field in enumerate(fields, 1):
                val = s.get(field)
                cell = ws.cell(row=row, column=col, value=val)
                data_style(cell, fill=row_fill)
            total_mt += s.get("quantity_mt") or 0
            total_bags += s.get("quantity_bags") or 0
            row += 1

        # Subtotal row
        subtotal_fill = PatternFill("solid", fgColor="D9E1F2")
        subtotal_label = ws.cell(row=row, column=1, value=f"SUBTOTAL — {origin}")
        subtotal_label.font = Font(bold=True)
        subtotal_label.fill = subtotal_fill
        subtotal_label.border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        mt_cell = ws.cell(row=row, column=3, value=total_mt if total_mt else None)
        mt_cell.font = Font(bold=True)
        mt_cell.fill = subtotal_fill
        mt_cell.border = BORDER

        bags_cell = ws.cell(row=row, column=4, value=total_bags if total_bags else None)
        bags_cell.font = Font(bold=True)
        bags_cell.fill = subtotal_fill
        bags_cell.border = BORDER

        for col in range(5, len(headers) + 1):
            ws.cell(row=row, column=col).fill = subtotal_fill
            ws.cell(row=row, column=col).border = BORDER

        row += 2  # blank line between origins

    auto_width(ws)

    # Grand total
    total_mt_all = sum(s.get("quantity_mt") or 0 for s in shipments)
    total_bags_all = sum(s.get("quantity_bags") or 0 for s in shipments)
    ws.cell(row=row, column=1, value=f"TOTAL — ALL ORIGINS  |  {total_mt_all:.0f} MT  |  {total_bags_all:.0f} bags").font = Font(bold=True, size=11)


# ── Tab 3: Pending Subjects ──────────────────────────────────────────────────

def write_pending_subjects(wb, subjects):
    ws = wb.create_sheet("Pending Subjects")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = ["Urgency", "Reference", "Subject Type", "Description", "Due Date", "Reported By", "Email Date"]
    fields = ["urgency", "reference", "subject_type", "description", "due_date", "reported_by", "source_email_date"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)

    for row_idx, s in enumerate(subjects, 2):
        urgency = s.get("urgency", "NORMAL")
        row_fill = RED_FILL if urgency == "HIGH" else ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col, field in enumerate(fields, 1):
            val = s.get(field)
            if field == "source_email_date" and val:
                val = val[:10]
            cell = ws.cell(row=row_idx, column=col, value=val)
            data_style(cell, fill=row_fill, bold=(urgency == "HIGH" and col == 1))

    auto_width(ws)


# ── Tab 4: Email Log ─────────────────────────────────────────────────────────

def write_email_log(wb, email_log):
    ws = wb.create_sheet("Email Log")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = ["Date", "From", "Subject", "Vessel Found", "BL Found", "Pending Subjects"]
    fields = ["date", "from", "subject", "vessel_found", "bl_found", "pending_subjects"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        header_style(cell)

    for row_idx, e in enumerate(email_log, 2):
        alt = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col, field in enumerate(fields, 1):
            val = e.get(field)
            if field == "date" and val:
                val = val[:10]
            if isinstance(val, list):
                val = ", ".join(val)
            cell = ws.cell(row=row_idx, column=col, value=val)
            data_style(cell, fill=alt)

    auto_width(ws)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not SHIPMENTS_FILE.exists():
        print(f"ERROR: {SHIPMENTS_FILE} not found. Run parse_shipments.py first.")
        sys.exit(1)

    shipments_data = json.loads(SHIPMENTS_FILE.read_text())
    changes_data = {}
    if CHANGES_FILE.exists():
        changes_data = json.loads(CHANGES_FILE.read_text())
    else:
        print("NOTE: changes.json not found — Changes tab will be empty.")
        changes_data = {"baseline": True, "note": "No comparison data available."}

    today = datetime.today().strftime("%Y-%m-%d")
    output_file = OUTPUT_DIR / f"ops_report_{today}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    shipments = shipments_data.get("active_shipments", [])
    shipments_lookup = {(s.get("vessel") or "") + "|" + (s.get("bl_number") or ""): s
                        for s in shipments}

    print("Writing Active Shipments tab...")
    write_active_shipments(wb, shipments)

    print("Writing By Origin tab...")
    write_by_origin(wb, shipments)

    print("Writing Changes This Week tab...")
    write_changes(wb, changes_data, shipments_lookup)

    print("Writing Pending Subjects tab...")
    write_pending_subjects(wb, shipments_data.get("pending_subjects", []))

    print("Writing Email Log tab...")
    write_email_log(wb, shipments_data.get("email_log", []))

    wb.save(str(output_file))
    print(f"\nReport saved → {output_file}")
    print(f"  Shipments:        {len(shipments)}")
    print(f"  Pending subjects: {len(shipments_data.get('pending_subjects', []))}")
    print(f"  Emails processed: {shipments_data.get('email_count', '?')}")

    return str(output_file)


if __name__ == "__main__":
    main()
